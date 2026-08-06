import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

def generate_jay_qa_excel():
    wb = openpyxl.Workbook()
    
    # Define Styles
    header_fill = PatternFill(start_color="4B638C", end_color="4B638C", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    title_font = Font(name="Calibri", size=16, bold=True, color="F55825")
    subtitle_font = Font(name="Calibri", size=11, italic=True, color="4B638C")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=10.5)
    
    p0_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Soft Red
    p0_font = Font(name="Calibri", size=10, bold=True, color="991B1B")
    
    p1_fill = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid") # Soft Orange
    p1_font = Font(name="Calibri", size=10, bold=True, color="9A3412")
    
    p2_fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid") # Soft Yellow
    p2_font = Font(name="Calibri", size=10, bold=True, color="854D0E")

    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    # -------------------------------------------------------------
    # TAB 1: Jay QA - 30 Jul 2026 (MAIN QA FEEDBACK LOG)
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Jay QA - 30 Jul 2026"
    ws1.views.sheetView[0].showGridLines = True
    
    # Header Banner
    ws1.merge_cells("A1:H1")
    ws1["A1"] = "Jay's Live Testing Feedback — 30 Jul 2026"
    ws1["A1"].font = title_font
    
    ws1.merge_cells("A2:H2")
    ws1["A2"] = "Platform URL: https://udenstoragelearningpath.z29.web.core.windows.net/ | Tested By: Vasanthi (Frontend Engineer) | Reviewed By: Jay Bhaskar (Product Manager)"
    ws1["A2"].font = subtitle_font
    
    headers = ["#", "Screen", "Issue / Tested Observation", "Evidence", "Recommended Fix / UX Action", "Priority", "Owner", "Status"]
    
    ws1.append([]) # Blank row 3
    ws1.append(headers) # Row 4
    
    for col_num, header in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    qa_data = [
        (1, "Dashboard - My Learning Paths", "Module Completion Tracking: Upon completing the 'Full Stack Web Developer - Carrier' module, progress bar updates to 100% and displays green 'Learning Completed' badge. Total Learning Paths metric updates to 76 and Completed LP updates to 1.", "Dashboard (/dashboard) - Path Library Card & Insights Panel", "Feature working cleanly. Recommend adding a celebratory confetti animation and certificate download button upon module completion to boost student motivation.", "P2 - Medium", "Vasanthi", "Tested by Vasanthi (30 Jul 2026)"),
        (2, "Learning Path - Quiz Assessment Modal", "Company-Specific Question Tagging: MCQ quiz displays 'Asked in: Meta, Amazon, Microsoft' badge below questions (e.g. Question 9: React Create App command). Makes it easy for students to connect quizzes to target company rounds.", "Quiz Modal (/learn) - MCQ Question 9 of 25", "Feature working cleanly. Recommend making company tags clickable so students can filter all practice questions asked specifically by Meta or Amazon.", "P1 - High", "Vasanthi", "Tested by Vasanthi (30 Jul 2026)"),
        (3, "Quiz Result - Skill Breakdown Scorecard", "Granular Skill Review & Improvement Insights: Scorecard displays overall score (36%), pass threshold (40 pts), and granular skill breakdown (React: 4/5 'good understanding', Node.js: 1/5 'need to improve', Agile: 1.5/5 'need to improve').", "Quiz Results Overlay (/learn) - Scorecard Modal", "Feature working cleanly. Recommend adding a 1-click 'Generate Targeted Remedial Study Plan' CTA button directly next to skills flagged as 'need to improve'.", "P1 - High", "Vasanthi", "Tested by Vasanthi (30 Jul 2026)"),
        (4, "Dashboard - Credits & Quiz Counter Sync", "Real-Time Credit & Counter Deduction: Header credits deduct dynamically per quiz attempt (AVAILABLE: 3753 -> 3752 -> 3751, CONSUMED: 2824 -> 2825 -> 2826) and Quiz Attempt Counter updates from 1 to 2 on module card.", "Dashboard Header & Path Library Badge", "System sync working correctly. Recommend adding a subtle toast notification when credits are consumed to make point ledger changes transparent to the learner.", "P2 - Medium", "Vasanthi", "Tested by Vasanthi (30 Jul 2026)"),
        (5, "Learning Path - Target Role Description", "Character Limit Validation (<100 Chars): Input textarea enforces 100 character minimum. Displays red error message 'Please enter at least 100 characters' (e.g. at 95/100 chars) and disables generation.", "Generator Page (/learn) - Describe your target role input", "Working cleanly. Recommend adding a dynamic character counter progress ring next to '95 / 100' so users visually see how many characters remain.", "P2 - Medium", "Vasanthi", "Tested by Vasanthi (30 Jul 2026)"),
        (6, "Learning Path - Review Acknowledgement", "Pre-Generation Review Modal: Upon submitting 100+ characters, system opens 'REVIEW ACKNOWLEDGEMENT' modal displaying About Company, Summarized Role, and Previously Asked Questions.", "Review Acknowledgement Overlay (/learn)", "Working cleanly. Recommend adding an 'Edit Prompt' button inside the modal to allow quick prompt adjustments before consenting to generation.", "P1 - High", "Vasanthi", "Tested by Vasanthi (30 Jul 2026)"),
        (7, "Learning Path - AI Generation Engine", "Real-Time Generation Toast & Streaming: Displays loading banner 'Please wait... The AI is thoughtfully shaping your journey' while streaming summarized role and skill nodes.", "Generation View (/learn) - Loading Banner Overlay", "Working cleanly. Recommend adding an estimated time remaining indicator (e.g. ~15 seconds) inside the loading banner for better feedback.", "P2 - Medium", "Vasanthi", "Tested by Vasanthi (30 Jul 2026)"),
        (8, "Learning Path - Generated Tree View", "Interactive Skill Nodes & Credit Consumption: Renders 'Frontend Engineering Learning Path Designer' with Skill 1 (Intro), Skill 2 (Curriculum Design), and sub-topics. Deducts 20 credits (AVAILABLE: 3751 -> 3731, CONSUMED: 2826 -> 2846).", "Generated Path View (/learn) - Skill Nodes & Header", "Working cleanly. Recommend making 'Mark as Completed' button trigger instant node checkmarks without full page reload.", "P1 - High", "Vasanthi", "Tested by Vasanthi (30 Jul 2026)"),
        (9, "Dashboard - Learning Path Synchronization", "Real-Time Dashboard & Metric Sync: Generated path instantly syncs to Dashboard. Total Learning Paths updates from 76 to 77, and new path appears at top of 'My Learning Paths' with 'Start Learning' button.", "Dashboard (/dashboard) - Path Library & Insights Panel", "Working cleanly. Tooltip 'This title is based on historical data' renders properly. System sync is 100% operational.", "P2 - Medium", "Vasanthi", "Tested by Vasanthi (30 Jul 2026)"),
        (10, "Career Acceleration - Guided Tour", "Guided Interactive Tour Overlay: Clicking '+ Take a Tour' button at bottom left launches interactive onboarding guide tooltips explaining search filters, resume parsing, and match feeds.", "Career Acceleration (/job-search) - Bottom Left Button", "Working cleanly. Onboarding tour guides students effectively through the platform features.", "P2 - Medium", "Vasanthi", "Tested by Vasanthi (30 Jul 2026)"),
        (11, "Career Acceleration - Job Bookmark / Save", "Job Save Toast & Counter Increment: Clicking 'Save' triggers toast 'Job saved successfully.', updates button to active 'Saved' badge, and increments Saved Jobs counter from 4 to 5 in top tab and left candidate sidebar.", "Matched Jobs Feed (/job-search) - Deloitte Digital Card", "Working cleanly. Real-time counter synchronization verified across candidate sidebar and header tabs.", "P2 - Medium", "Vasanthi", "Tested by Vasanthi (30 Jul 2026)"),
        (12, "Career Acceleration - Job View Details Drawer", "Role & Requirements Breakdown: Clicking 'View details' opens slide-out drawer rendering About Role, Requirements, What You'll Do, Skills Required, and Fit Scorecard (Skills Matched vs Skill Gaps).", "View Details Drawer (/job-search) - Goodway Group Role", "Working cleanly. Clear visual distinction between matched skills (green checkmarks) and skill gaps (orange warning pills).", "P1 - High", "Vasanthi", "Tested by Vasanthi (30 Jul 2026)"),
        (13, "Career Acceleration - Interview Rounds Breakdown", "Interview Preparation Insights: View Details drawer features 'INTERVIEW ROUNDS' section detailing Round 1 (Resume Screening 20-30 mins) and Round 2 (Online Coding Assessment 60-90 mins) with actionable prep tips.", "View Details Drawer (/job-search) - Interview Rounds Section", "Working cleanly. Gives candidates a structured preparation roadmap for company-specific interview rounds.", "P1 - High", "Vasanthi", "Tested by Vasanthi (30 Jul 2026)"),
        (14, "Career Acceleration - External Job Redirection", "External ATS Redirection: Clicking 'Apply now' inside job drawer seamlessly opens external company board (e.g. Goodway Group Greenhouse portal job-boards.greenhouse.io) in a new browser tab.", "External Tab Redirection (Greenhouse ATS Board)", "Working cleanly. Allows students to apply directly on official employer ATS portals.", "P1 - High", "Vasanthi", "Tested by Vasanthi (30 Jul 2026)"),
        (15, "Career Acceleration - 'Build Learning Path' Integration", "Auto-Populate Learning Path Redirection: Clicking 'Build learning path' inside job drawer redirects to /learn and automatically pre-populates target job description (293/100 chars) into generator input.", "Drawer Footer CTA -> Learning Path Generator (/learn)", "Working cleanly. Excellent seamless bridge between job discovery and custom AI learning path generation.", "P0 - Blocker", "Vasanthi", "Tested by Vasanthi (30 Jul 2026)"),
        (16, "Application Tracker - Kanban Board & Status View", "Application Stage Management: Application Tracker (/application-tracker) displays Kanban columns (Applied: 2, Screening: 1, Interview: 0, Offer: 0). Candidate can view application status updated by Admin/Employer.", "Application Tracker (/application-tracker) - Kanban View", "Working cleanly. Student can view live tracking status (e.g. 23d 14h ago) and edit submission records.", "P1 - High", "Vasanthi", "Tested by Vasanthi (30 Jul 2026)"),
        (17, "Application Tracker - View Mode Switching", "Multi-View Layout Controls (Kanban, Progress, Archived): Toggling 'Progress View' displays linear stage timeline (Applied 2 -> Screening 1 -> Interview 0 -> Offer 0) with clear progress indicators.", "Application Tracker (/application-tracker) - View Controls", "Working cleanly. Provides versatile tracking views for students monitoring multiple job applications.", "P2 - Medium", "Vasanthi", "Tested by Vasanthi (30 Jul 2026)"),
        (18, "Application Tracker - Saved Jobs Slide-Out Drawer", "Saved Jobs Quick Access Drawer: Clicking 'Saved' button on Application Tracker header opens slide-out drawer rendering all 5 Saved Jobs with fit scores, company tags, and instant CTAs (Quick view, View details, Build LP).", "Application Tracker (/application-tracker) - Saved Jobs Drawer", "Working cleanly. Allows candidates to review saved jobs and build custom learning paths without navigating away from the tracker.", "P1 - High", "Vasanthi", "Tested by Vasanthi (30 Jul 2026)"),
        (19, "Settings - Theme Customization Drawer", "Theme State Sync Bug (Semi Dark & Palette Swatches): Toggling 'Semi Dark' mode or color swatches only partially updates the left sidebar to dark navy while leaving main page container backgrounds in light mode, causing text contrast mismatch.", "Settings Customization Drawer (/application-tracker) - Theme: Semi Dark", "BUG OBSERVED. Unify theme state manager (ThemeContext / CSS :root variables) so toggling Lite vs Semi Dark updates all body, container, card, and sidebar background tokens globally in sync.", "P1 - High", "Vasanthi", "Tested by Vasanthi (30 Jul 2026)"),
        (20, "Admin Nexus - Live Metrics & Manage Jobs List", "Admin Job Management Overview: Admin Nexus (/job-management) displays live metric cards (Total Jobs: 8260, Admin Posted: 5, AI Discovered: 7953, Application Tracker: 3) and data list of open positions with skill tags.", "Admin Nexus (/job-management) - Live Metrics & Manage Jobs", "Working cleanly. Filtering by Status (Active/Disabled), Source (Admin Posted/AI Discovered), and Visibility works as expected.", "P1 - High", "Vasanthi", "Tested by Vasanthi (30 Jul 2026)"),
        (21, "Admin Nexus - 'Import Job from URL' Crawler", "AI URL Job Extractor Modal: Clicking 'Import from URL' opens modal parser. Validates robots.txt rules, auto-fills job title, company name, required skills, and salary from external posting URLs.", "Import Job from URL Modal (/job-management)", "Working cleanly. Streamlines enterprise job ingestion directly from career portals.", "P1 - High", "Vasanthi", "Tested by Vasanthi (30 Jul 2026)"),
        (22, "Admin Nexus - URL Crawler SPA Error Handling", "SPA Client-Rendered URL Handling: When crawling JavaScript/SPA-rendered URLs (e.g. EY Careers), system accurately catches non-accessible HTML and informs admin: 'Page content is loaded entirely by JavaScript. Try an ATS link.'", "Import Job from URL Modal Error Box (/job-management)", "Working cleanly. Prevents corrupted or empty job entries from being created in microservices DB. Recommend adding fallback manual text paste option inside error box.", "P2 - Medium", "Vasanthi", "Tested by Vasanthi (30 Jul 2026)"),
        (23, "Admin Nexus - Activity Monitor & Live Stream", "Real-Time Admin Telemetry & Engagement Logs: Activity Monitor tab displays Click-to-Apply (3.2%), Engagement (3%), Retention (3%), Live Event Stream log of candidate actions (e.g. 'shoaib applied to...'), and Hot Opportunities ranking.", "Admin Nexus (/job-management) - Activity Monitor Tab", "Working cleanly. Provides administrators with live telemetry on candidate job applications and engagement metrics.", "P1 - High", "Vasanthi", "Tested by Vasanthi (30 Jul 2026)")
    ]

    for row_idx, data in enumerate(qa_data, start=5):
        ws1.append(data)
        for col_idx in range(1, 9):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.font = normal_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            # Format # column
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center", vertical="top")
                
            # Format Priority Column
            if col_idx == 6:
                cell.alignment = Alignment(horizontal="center", vertical="top")
                val = str(cell.value)
                if "P0" in val:
                    cell.fill = p0_fill
                    cell.font = p0_font
                elif "P1" in val:
                    cell.fill = p1_fill
                    cell.font = p1_font
                elif "P2" in val:
                    cell.fill = p2_fill
                    cell.font = p2_font

    # Set Column Widths for WS1
    col_widths = {1: 6, 2: 24, 3: 40, 4: 28, 5: 42, 6: 14, 7: 10, 8: 24}
    for col_idx, width in col_widths.items():
        ws1.column_dimensions[get_column_letter(col_idx)].width = width

    # -------------------------------------------------------------
    # TAB 2: Summary (EXEC SUMMARY & MATCHED JOBS ANSWER)
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Summary")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2["A1"] = "UDEN Learning Path Platform — Live Testing Executive Summary"
    ws2["A1"].font = title_font
    
    ws2.append([])
    ws2.append(["PRODUCT MANAGER QUESTION ANSWER:"])
    ws2.cell(row=3, column=1).font = bold_font
    ws2.append(["Question:", "How many jobs you got after putting your resume?"])
    ws2.append(["Answer:", "After uploading candidate resume for identity shoaib (shoaib@noventiqai.com), the AI match engine dynamically calculated fit scores and populated 3 Matched Jobs directly under 'Matched for you':"])
    ws2.append(["", "1. Senior React & Frontend Engineer — 90% Fit Score (Admin Posted • Deloitte Digital • Bangalore)"])
    ws2.append(["", "2. Frontend Web Developer [three.js] — 71% Fit Score (AI Discovered • Valerie Group • India)"])
    ws2.append(["", "3. Global Full Stack Software Engineer — 71% Fit Score (AI Discovered • Goodway Group • Remote)"])
    ws2.append(["", "Platform Totals: 5 Saved Jobs, 2 Applications Tracked, 5 Admin Posted Jobs, 7,953 AI Discovered Jobs, 8,260 Total Platform Jobs."])
    
    ws2.append([])
    ws2.append(["TESTING MILESTONE SUMMARY MATRIX:"])
    ws2.cell(row=9, column=1).font = bold_font
    
    sum_headers = ["Module Area", "Checkpoints Tested", "Status", "Key Findings & Performance"]
    ws2.append(sum_headers)
    for c_idx in range(1, 5):
        cell = ws2.cell(row=10, column=c_idx)
        cell.fill = header_fill
        cell.font = header_font
        
    sum_rows = [
        ("Candidate Onboarding & Resume Parser", "5 Checkpoints", "PASSED", "Parsed 19 skills seamlessly; experience level calculated at 3.0 yrs."),
        ("AI Learning Path Generator & Quizzes", "5 Checkpoints", "PASSED", "100-char validation, Review Acknowledgement, 20-credit deduction, 77 total paths."),
        ("Career Acceleration & ATS Redirection", "6 Checkpoints", "PASSED", "90% fit score for Deloitte job, Greenhouse ATS redirection, Build LP auto-populate."),
        ("Application Tracker & Saved Drawer", "3 Checkpoints", "PASSED", "Kanban view (Applied: 2, Screening: 1), Progress View timeline, Saved Jobs drawer (5 saved)."),
        ("Settings Theme Drawer", "1 Checkpoint", "FLAGGED BUG", "Item 19: Semi Dark mode causes partial theme sync (dark sidebar, light content area)."),
        ("Admin Nexus & URL Crawler", "3 Checkpoints", "PASSED", "Tracks 8,260 jobs, validates robots.txt, catches SPA JS-rendered URLs, streams live logs.")
    ]
    for r_data in sum_rows:
        ws2.append(r_data)
        
    for col_i in range(1, 5):
        ws2.column_dimensions[get_column_letter(col_i)].width = 30
    ws2.column_dimensions["D"].width = 65

    # -------------------------------------------------------------
    # TAB 3: UI-QA Bug Fixes
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="UI-QA Bug Fixes")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3["A1"] = "UDEN Platform — UI/UX Recommended Fixes & Bug Tracking"
    ws3["A1"].font = title_font
    ws3.append([])
    
    bug_headers = ["Bug / Fix ID", "Component / Screen", "Observed Defect", "Recommended Technical Fix", "Priority"]
    ws3.append(bug_headers)
    for c_idx in range(1, 6):
        cell = ws3.cell(row=3, column=c_idx)
        cell.fill = header_fill
        cell.font = header_font
        
    bugs_data = [
        ("BUG-01", "Settings Theme Drawer", "Item 19: Toggling Semi Dark mode updates sidebar to dark navy but leaves content area light, causing contrast mismatch.", "Unify theme state manager (ThemeContext / CSS :root variables) to update all background tokens globally in sync.", "P1 - High"),
        ("BUG-02", "SPA Routing Infrastructure", "Reloading direct URLs (/job-search, /job-management) triggers Azure Blob Storage 404 error.", "Add Azure CDN URL rewrite rule mapping sub-routes back to index.html.", "P1 - High"),
        ("UI-01", "Candidate Header Navigation", "Credit pills (AVAILABLE - 3731, CONSUMED - 2846) use default blue styling.", "Standardize credit pills with soft Steel Blue outlines (border: 1px solid rgba(75, 99, 140, 0.25)) and map avatar to Steel Blue (#4B638C).", "P2 - Medium"),
        ("UI-02", "Candidate Sidebar Skill Chips", "Skill chips wrap tightly against experience metrics in profile card.", "Add 6px flex gap spacing and soft gold/steel blue chip styling (background: #FFFDF0, color: #4B638C, border: 1.5px solid #F7BC08).", "P2 - Medium"),
        ("UI-03", "Matched Job Card Fit Badge", "The 90% fit badge renders as a plain text pill.", "Upgrade 90% fit badge to a Sunburst Gold pill (#F7BC08) with a sparkle icon. Map primary action buttons to Coral Orange (#F55825).", "P2 - Medium"),
        ("UI-04", "Admin Nexus Data Table", "In /job-management, company tags sit tightly against status tags.", "Add 12px margin spacing between tags and highlight + Add New Job CTA button with a Primary Coral Orange (#F55825) gradient fill.", "P2 - Medium"),
        ("UI-05", "Mobile Viewport Responsiveness", "Filter panel on left sidebar causes horizontal overflow on mobile screens (<768px).", "Convert static filter column into a touch-optimized collapsible drawer toggle for mobile viewports.", "P2 - Medium")
    ]
    for b_data in bugs_data:
        ws3.append(b_data)
        
    for col_i in range(1, 6):
        ws3.column_dimensions[get_column_letter(col_i)].width = 25
    ws3.column_dimensions["C"].width = 45
    ws3.column_dimensions["D"].width = 50

    # Save to public and scratch
    out_public = r"c:\Users\vasan\.gemini\antigravity\scratch\uden-app-react\public\Jay_QA_Live_Testing_Feedback_30_Jul_2026.xlsx"
    wb.save(out_public)
    print(f"SAVED_XLSX: {out_public}")

if __name__ == "__main__":
    generate_jay_qa_excel()
